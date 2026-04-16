from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, F, Case, When, DecimalField
from django.http import HttpResponse, HttpResponseForbidden
from datetime import datetime, time
from django.utils import timezone
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors

# Import permission mixins
try:
    from accounts.permissions import (
        SorViewPermissionMixin, SorAddPermissionMixin, 
        SorEditPermissionMixin, SorDeletePermissionMixin
    )
except ImportError:
    # Fallback mixins if permissions not available
    from django.contrib.auth.mixins import LoginRequiredMixin
    class SorViewPermissionMixin(LoginRequiredMixin):
        pass
    class SorAddPermissionMixin(LoginRequiredMixin):
        pass
    class SorEditPermissionMixin(LoginRequiredMixin):
        pass
    class SorDeletePermissionMixin(LoginRequiredMixin):
        pass

from .models import SOR
from .forms import SORForm, SORFilterForm
from .notification import SORNotification
from django.contrib.auth import get_user_model
from trips.models import Trip
from vehicles.models import Vehicle
User = get_user_model()

@login_required
def sor_create(request):
    # Check permission using the new permission system
    if not request.user.has_module_permission('sor', 'add'):
        messages.error(request, 'You do not have permission to create SOR entries.')
        return redirect('sor_list')
    if request.method == 'POST':
        post_data = request.POST.copy()
        # If "Others" is selected, replace with the typed value
        if post_data.get('from_location') == 'Others':
            other = post_data.get('from_location_other', '').strip()
            if other:
                post_data['from_location'] = other
        if post_data.get('to_location') == 'Others':
            other = post_data.get('to_location_other', '').strip()
            if other:
                post_data['to_location'] = other
        form = SORForm(post_data)
        if form.is_valid():
            sor = form.save(commit=False)
            sor.created_by = request.user

            if sor.source_type == 'outsourced_manual':
                sor.status = 'completed'
                if sor.start_odometer is not None and sor.end_odometer is not None:
                    sor.distance_km = sor.end_odometer - sor.start_odometer

            sor.save()
            if sor.source_type == 'company' and sor.driver:
                # Create notification for driver only for regular company flow.
                SORNotification.objects.create(
                    sor=sor,
                    driver=sor.driver,
                    message=f"You have a new SOR assignment from {sor.from_location} to {sor.to_location}. Please accept or reject.",
                )
                messages.success(request, 'SOR entry created and driver notified.')
            else:
                messages.success(request, 'Outsourced SOR entry created successfully.')
            return redirect('sor_list')
    else:
        source_type = request.GET.get('source_type', 'company')
        form = SORForm(initial={'source_type': source_type})
    return render(request, 'sor/sor_form.html', {'form': form})

@login_required
def sor_list(request):
    # Check permission to view SOR entries
    if not request.user.has_module_permission('sor', 'view'):
        messages.error(request, 'You do not have permission to view SOR entries.')
        return redirect('dashboard')
    
    # Filter SORs based on user type (not permissions)
    if request.user.user_type in ['admin', 'manager', 'vehicle_manager', 'sor_head']:
        # Management users and SOR Head can see all SORs
        sors = SOR.objects.all()
    elif request.user.user_type == 'driver':
        # Drivers can only see SORs where they are the assigned driver
        sors = SOR.objects.filter(driver=request.user)
    else:
        # Other user types (sor_team, generator_user) see SORs they created
        sors = SOR.objects.filter(created_by=request.user)
    
    # Initialize filter form
    filter_form = SORFilterForm(request.GET or None)
    
    # Apply filters if form is valid
    if filter_form.is_valid():
        # Search filter
        search = filter_form.cleaned_data.get('search')
        if search:
            sors = sors.filter(
                Q(id__icontains=search) |
                Q(goods_value__icontains=search) |
                Q(from_location__icontains=search) |
                Q(to_location__icontains=search) |
                Q(outsourced_vehicle_text__icontains=search) |
                Q(outsourced_driver_text__icontains=search) |
                Q(vendor_name__icontains=search) |
                Q(driver__first_name__icontains=search) |
                Q(driver__last_name__icontains=search) |
                Q(vehicle__license_plate__icontains=search)
            )
        
        # Status filter
        status = filter_form.cleaned_data.get('status')
        if status:
            sors = sors.filter(status=status)

        source_type = filter_form.cleaned_data.get('source_type')
        if source_type:
            sors = sors.filter(source_type=source_type)
        
        # From location filter
        from_location = filter_form.cleaned_data.get('from_location')
        if from_location:
            sors = sors.filter(from_location=from_location)
        
        # To location filter
        to_location = filter_form.cleaned_data.get('to_location')
        if to_location:
            sors = sors.filter(to_location=to_location)
        
        # Vehicle filter
        vehicle = filter_form.cleaned_data.get('vehicle')
        if vehicle:
            sors = sors.filter(vehicle=vehicle)
        
        # Driver filter
        driver = filter_form.cleaned_data.get('driver')
        if driver:
            sors = sors.filter(driver=driver)
        
        # Date range filters
        date_from = filter_form.cleaned_data.get('date_from')
        if date_from:
            # Filter from start of the selected date
            start_datetime = timezone.make_aware(datetime.combine(date_from, time.min))
            sors = sors.filter(created_at__gte=start_datetime)
        
        date_to = filter_form.cleaned_data.get('date_to')
        if date_to:
            # Filter until end of the selected date
            end_datetime = timezone.make_aware(datetime.combine(date_to, time.max))
            sors = sors.filter(created_at__lte=end_datetime)
    
    # Handle sorting
    sort_by = request.GET.get('sort', 'created_at')
    order = request.GET.get('order', 'desc')
    
    # Define allowed sort fields
    allowed_sort_fields = {
        'id': 'id',
        'source_type': 'source_type',
        'goods_value': 'goods_value',
        'created_at': 'created_at',
        'from_location': 'from_location',
        'to_location': 'to_location',
        'vehicle': 'vehicle__license_plate',
        'rate_per_km': 'vehicle__rate_per_km',
        'distance_km': 'distance_km',
        'driver': 'driver__first_name',
        'status': 'status',
    }
    
    # For computed fields, we need to annotate the queryset
    if sort_by in ['transport_cost', 'transport_cost_percentage']:
        # Annotate with computed transport cost
        sors = sors.annotate(
            computed_transport_cost=Case(
                When(distance_km__isnull=False, vehicle__rate_per_km__isnull=False,
                     then=F('distance_km') * F('vehicle__rate_per_km')),
                default=None,
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )
        
        if sort_by == 'transport_cost':
            sort_field = 'computed_transport_cost'
        elif sort_by == 'transport_cost_percentage':
            # Annotate with computed transport cost percentage
            sors = sors.annotate(
                computed_transport_cost_percentage=Case(
                    When(computed_transport_cost__isnull=False, goods_value__gt=0,
                         then=(F('computed_transport_cost') / F('goods_value')) * 100),
                    default=None,
                    output_field=DecimalField(max_digits=8, decimal_places=2)
                )
            )
            sort_field = 'computed_transport_cost_percentage'
            
        if order == 'desc':
            sort_field = '-' + sort_field
        sors = sors.order_by(sort_field)
        
    # Validate sort field for regular database fields
    elif sort_by in allowed_sort_fields:
        sort_field = allowed_sort_fields[sort_by]
        if order == 'desc':
            sort_field = '-' + sort_field
        sors = sors.order_by(sort_field)
    else:
        # Default ordering
        sors = sors.order_by('-created_at')

    # Pagination - 30 items per page
    paginator = Paginator(sors, 30)
    page = request.GET.get('page')
    
    try:
        sors_page = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        sors_page = paginator.page(1)
    except EmptyPage:
        # If page is out of range, deliver last page
        sors_page = paginator.page(paginator.num_pages)
    
    context = {
        'sors': sors_page,
        'filter_form': filter_form,
        'total_count': paginator.count,
        'has_filters': any(filter_form.cleaned_data.values()) if filter_form.is_valid() else False,
        'current_sort': sort_by,
        'current_order': order,
    }
    
    return render(request, 'sor/sor_list.html', context)

# --- SOR View, Edit, Delete Placeholder Views ---
from django.http import HttpResponseForbidden

@login_required
def sor_view(request, pk):
    sor = get_object_or_404(SOR, pk=pk)
    
    # Check permission to view SOR entries
    if not request.user.has_module_permission('sor', 'view'):
        return HttpResponseForbidden('You do not have permission to view this SOR.')
    
    # Check if user can view this specific SOR (based on user type)
    if request.user.user_type in ['admin', 'manager', 'vehicle_manager', 'sor_head']:
        # Management users and SOR Head can view any SOR
        pass
    elif request.user.user_type == 'driver':
        # Drivers can only view SORs where they are the assigned driver
        if sor.driver != request.user:
            return HttpResponseForbidden('You can only view SORs where you are the assigned driver.')
    else:
        # Other user types can only view SORs they created
        if sor.created_by != request.user:
            return HttpResponseForbidden('You can only view SORs you created.')
    
    return render(request, 'sor/sor_form.html', {'form': None, 'sor': sor, 'view_only': True})

@login_required
def sor_edit(request, pk):
    sor = get_object_or_404(SOR, pk=pk)
    # Check permission using the new permission system
    if not request.user.has_module_permission('sor', 'edit'):
        return HttpResponseForbidden('You do not have permission to edit this SOR.')
    
    # Check if user can edit this specific SOR (based on user type)
    if request.user.user_type in ['admin', 'manager', 'vehicle_manager', 'sor_head']:
        # Management users and SOR Head can edit any SOR
        pass
    elif request.user.user_type == 'driver':
        # Drivers can only edit SORs where they are the assigned driver
        if sor.driver != request.user:
            return HttpResponseForbidden('You can only edit SORs where you are the assigned driver.')
    else:
        # Other user types can only edit SORs they created
        if sor.created_by != request.user:
            return HttpResponseForbidden('You can only edit SORs you created.')
    
    if request.method == 'POST':
        form = SORForm(request.POST, instance=sor)
        if form.is_valid():
            updated_sor = form.save()
            if updated_sor.source_type == 'outsourced_manual':
                updated_sor.status = 'completed'
                if updated_sor.start_odometer is not None and updated_sor.end_odometer is not None:
                    updated_sor.distance_km = updated_sor.end_odometer - updated_sor.start_odometer
                updated_sor.save(update_fields=['status', 'distance_km'])
                messages.success(request, 'Outsourced SOR entry updated successfully.')
            else:
                # Update or create notification for driver
                from .notification import SORNotification
                if updated_sor.driver:
                    notif_message = f"SOR assignment updated: {updated_sor.from_location} to {updated_sor.to_location}. Please check details."
                    notif_qs = SORNotification.objects.filter(sor=updated_sor, driver=updated_sor.driver, is_read=False).order_by('-created_at')
                    if notif_qs.exists():
                        notif = notif_qs.first()
                        notif.message = notif_message
                        notif.save()
                    else:
                        SORNotification.objects.create(
                            sor=updated_sor,
                            driver=updated_sor.driver,
                            message=notif_message
                        )
                messages.success(request, 'SOR entry updated successfully and driver notified.')
            return redirect('sor_list')
    else:
        form = SORForm(instance=sor)
    return render(request, 'sor/sor_form.html', {'form': form, 'sor': sor, 'edit_mode': True})

@login_required
def sor_delete(request, pk):
    sor = get_object_or_404(SOR, pk=pk)
    # Check permission using the new permission system
    if not request.user.has_module_permission('sor', 'delete'):
        return HttpResponseForbidden('You do not have permission to delete this SOR.')
    
    # Check if user can delete this specific SOR (based on user type)
    if request.user.user_type in ['admin', 'manager', 'vehicle_manager', 'sor_head']:
        # Management users and SOR Head can delete any SOR
        pass
    elif request.user.user_type == 'driver':
        # Drivers can only delete SORs where they are the assigned driver
        if sor.driver != request.user:
            return HttpResponseForbidden('You can only delete SORs where you are the assigned driver.')
    else:
        # Other user types can only delete SORs they created
        if sor.created_by != request.user:
            return HttpResponseForbidden('You can only delete SORs you created.')
    
    if request.method == 'POST':
        sor.delete()
        messages.success(request, 'SOR entry deleted.')
        return redirect('sor_list')
    return render(request, 'sor/sor_confirm_delete.html', {'sor': sor})

@login_required
def sor_accept(request, pk):
    sor = get_object_or_404(SOR, pk=pk, driver=request.user)
    if sor.source_type != 'company':
        messages.error(request, 'Outsourced SOR entries do not support driver accept/start workflow.')
        return redirect('sor_list')
    if sor.status == 'pending':
        sor.status = 'driver_accepted'
        sor.save()
        # Start trip automatically
        # Use SOR's from_location, to_location, vehicle, driver, and distance
        start_odometer = sor.vehicle.current_odometer or 0
        trip = Trip.objects.create(
            vehicle=sor.vehicle,
            driver=sor.driver,
            start_time=timezone.now(),
            start_odometer=start_odometer,
            origin=sor.from_location,
            destination=sor.to_location,
            purpose=f"SOR Goods Value: {sor.goods_value}",
            notes=f"Started from SOR entry #{sor.id}",
            status='ongoing',
            entry_type='real_time',
        )
        sor.trip = trip
        sor.status = 'in_progress'
        sor.save()
        # Mark notification as read
        SORNotification.objects.filter(sor=sor, driver=request.user, is_read=False).update(is_read=True)
        messages.success(request, 'SOR accepted and trip started.')
    return redirect('sor_list')

@login_required
def sor_export(request):
    """Export SOR data with current filters applied"""
    export_format = request.GET.get('format', 'csv').lower()
    
    # Filter SORs based on user type (same logic as sor_list)
    if request.user.user_type in ['admin', 'manager', 'vehicle_manager', 'sor_head']:
        # Management users and SOR Head can export all SORs
        sors = SOR.objects.all()
    elif request.user.user_type == 'driver':
        # Drivers can only export SORs where they are the assigned driver
        sors = SOR.objects.filter(driver=request.user)
    else:
        # Other user types can only export SORs they created
        sors = SOR.objects.filter(created_by=request.user)
    
    # Apply the same filters as in sor_list view
    filter_form = SORFilterForm(request.GET or None)
    
    if filter_form.is_valid():
        # Search filter
        search = filter_form.cleaned_data.get('search')
        if search:
            sors = sors.filter(
                Q(id__icontains=search) |
                Q(goods_value__icontains=search) |
                Q(from_location__icontains=search) |
                Q(to_location__icontains=search) |
                Q(outsourced_vehicle_text__icontains=search) |
                Q(outsourced_driver_text__icontains=search) |
                Q(vendor_name__icontains=search) |
                Q(driver__first_name__icontains=search) |
                Q(driver__last_name__icontains=search) |
                Q(vehicle__license_plate__icontains=search)
            )
        
        # Status filter
        status = filter_form.cleaned_data.get('status')
        if status:
            sors = sors.filter(status=status)

        source_type = filter_form.cleaned_data.get('source_type')
        if source_type:
            sors = sors.filter(source_type=source_type)
        
        # From location filter
        from_location = filter_form.cleaned_data.get('from_location')
        if from_location:
            sors = sors.filter(from_location=from_location)
        
        # To location filter
        to_location = filter_form.cleaned_data.get('to_location')
        if to_location:
            sors = sors.filter(to_location=to_location)
        
        # Vehicle filter
        vehicle = filter_form.cleaned_data.get('vehicle')
        if vehicle:
            sors = sors.filter(vehicle=vehicle)
        
        # Driver filter
        driver = filter_form.cleaned_data.get('driver')
        if driver:
            sors = sors.filter(driver=driver)
        
        # Date range filters
        date_from = filter_form.cleaned_data.get('date_from')
        if date_from:
            start_datetime = timezone.make_aware(datetime.combine(date_from, time.min))
            sors = sors.filter(created_at__gte=start_datetime)
        
        date_to = filter_form.cleaned_data.get('date_to')
        if date_to:
            end_datetime = timezone.make_aware(datetime.combine(date_to, time.max))
            sors = sors.filter(created_at__lte=end_datetime)
    
    # Handle sorting (same as sor_list)
    sort_by = request.GET.get('sort', 'created_at')
    order = request.GET.get('order', 'desc')
    
    allowed_sort_fields = {
        'id': 'id',
        'source_type': 'source_type',
        'goods_value': 'goods_value',
        'created_at': 'created_at',
        'from_location': 'from_location',
        'to_location': 'to_location',
        'vehicle': 'vehicle__license_plate',
        'rate_per_km': 'vehicle__rate_per_km',
        'distance_km': 'distance_km',
        'driver': 'driver__first_name',
        'status': 'status',
    }
    
    # For computed fields, annotate the queryset
    if sort_by in ['transport_cost', 'transport_cost_percentage']:
        sors = sors.annotate(
            computed_transport_cost=Case(
                When(distance_km__isnull=False, vehicle__rate_per_km__isnull=False,
                     then=F('distance_km') * F('vehicle__rate_per_km')),
                default=None,
                output_field=DecimalField(max_digits=12, decimal_places=2)
            )
        )
        
        if sort_by == 'transport_cost':
            sort_field = 'computed_transport_cost'
        elif sort_by == 'transport_cost_percentage':
            sors = sors.annotate(
                computed_transport_cost_percentage=Case(
                    When(computed_transport_cost__isnull=False, goods_value__gt=0,
                         then=(F('computed_transport_cost') / F('goods_value')) * 100),
                    default=None,
                    output_field=DecimalField(max_digits=8, decimal_places=2)
                )
            )
            sort_field = 'computed_transport_cost_percentage'
            
        if order == 'desc':
            sort_field = '-' + sort_field
        sors = sors.order_by(sort_field)
        
    elif sort_by in allowed_sort_fields:
        sort_field = allowed_sort_fields[sort_by]
        if order == 'desc':
            sort_field = '-' + sort_field
        sors = sors.order_by(sort_field)
    else:
        sors = sors.order_by('-created_at')
    
    # Ensure export uses SOR ID ascending order regardless of current sort
    # Get all SOR data for export (no pagination)
    sors_data = sors.order_by('id').select_related('vehicle', 'driver', 'created_by')
    
    if export_format == 'csv':
        return _export_csv(sors_data)
    elif export_format == 'excel':
        return _export_excel(sors_data)
    elif export_format == 'pdf':
        return _export_pdf(sors_data)
    else:
        messages.error(request, 'Invalid export format.')
        return redirect('sor_list')

def _export_csv(sors_data):
    """Export SOR data as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="sor_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    headers = [
        'Sr. No.', 'SOR ID', 'Type', 'Goods Value', 'Created By', 'Created At', 'From Location', 'To Location',
        'Vehicle', 'Rate per KM', 'Distance (km)', 'Transport Cost', 'Transport % of Goods Value',
        'Driver', 'Status'
    ]
    writer.writerow(headers)
    
    # Write data with serial numbers
    for index, sor in enumerate(sors_data, 1):
        transport_cost = ''
        transport_percentage = ''
        
        if sor.distance_km and sor.vehicle and sor.vehicle.rate_per_km:
            transport_cost = f"{sor.distance_km * sor.vehicle.rate_per_km:.2f}"
            if sor.goods_value and sor.goods_value > 0:
                transport_percentage = f"{(sor.distance_km * sor.vehicle.rate_per_km / sor.goods_value * 100):.2f}%"
        
        row = [
            index,  # Serial number
            sor.id,  # Original SOR ID
            sor.get_source_type_display(),
            sor.goods_value,
            sor.created_by.get_full_name() if sor.created_by else '--',
            sor.created_at.strftime('%d %b %Y, %H:%M') if sor.created_at else '--',
            sor.from_location,
            sor.to_location,
            str(sor.vehicle) if sor.vehicle else (sor.outsourced_vehicle_text or '--'),
            sor.vehicle.rate_per_km if sor.vehicle and sor.vehicle.rate_per_km is not None else (sor.outsourced_rate_per_km if sor.outsourced_rate_per_km is not None else '--'),
            f"{sor.distance_km:.2f}" if sor.distance_km else '--',
            transport_cost or '--',
            transport_percentage or '--',
            str(sor.driver) if sor.driver else (sor.outsourced_driver_text or '--'),
            sor.get_status_display()
        ]
        writer.writerow(row)
    
    return response

def _export_excel(sors_data):
    """Export SOR data as Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sor_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'SOR Export'
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = [
        'Sr. No.', 'SOR ID', 'Type', 'Goods Value', 'Created By', 'Created At', 'From Location', 'To Location',
        'Vehicle', 'Rate per KM', 'Distance (km)', 'Transport Cost', 'Transport % of Goods Value',
        'Driver', 'Status'
    ]
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data with serial numbers
    for row_num, sor in enumerate(sors_data, 2):
        transport_cost = ''
        transport_percentage = ''
        
        if sor.distance_km and sor.vehicle and sor.vehicle.rate_per_km:
            transport_cost = f"{sor.distance_km * sor.vehicle.rate_per_km:.2f}"
            if sor.goods_value and sor.goods_value > 0:
                transport_percentage = f"{(sor.distance_km * sor.vehicle.rate_per_km / sor.goods_value * 100):.2f}%"
        
        data = [
            row_num - 1,  # Serial number (row_num starts from 2, so subtract 1)
            sor.id,  # Original SOR ID
            sor.get_source_type_display(),
            sor.goods_value,
            sor.created_by.get_full_name() if sor.created_by else '--',
            sor.created_at.strftime('%d %b %Y, %H:%M') if sor.created_at else '--',
            sor.from_location,
            sor.to_location,
            str(sor.vehicle) if sor.vehicle else (sor.outsourced_vehicle_text or '--'),
            sor.vehicle.rate_per_km if sor.vehicle and sor.vehicle.rate_per_km is not None else (sor.outsourced_rate_per_km if sor.outsourced_rate_per_km is not None else '--'),
            f"{sor.distance_km:.2f}" if sor.distance_km else '--',
            transport_cost or '--',
            transport_percentage or '--',
            str(sor.driver) if sor.driver else (sor.outsourced_driver_text or '--'),
            sor.get_status_display()
        ]
        
        for col, value in enumerate(data, 1):
            worksheet.cell(row=row_num, column=col, value=value)
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(response)
    return response

def _export_pdf(sors_data):
    """Export SOR data as PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sor_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter), topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("SOR Export Report", styles['Title'])
    elements.append(title)
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y at %H:%M')}", styles['Normal']))
    elements.append(Paragraph("<br/><br/>", styles['Normal']))
    
    # Prepare table data
    headers = [
        'Sr.', 'SOR ID', 'Type', 'Goods Value', 'Created At', 'From', 'To',
        'Vehicle', 'Rate/KM', 'Distance', 'Transport Cost', 'Transport %', 'Driver', 'Status'
    ]
    
    table_data = [headers]
    
    for index, sor in enumerate(sors_data, 1):
        transport_cost = ''
        transport_percentage = ''
        
        if sor.distance_km and sor.vehicle and sor.vehicle.rate_per_km:
            transport_cost = f"{sor.distance_km * sor.vehicle.rate_per_km:.2f}"
            if sor.goods_value and sor.goods_value > 0:
                transport_percentage = f"{(sor.distance_km * sor.vehicle.rate_per_km / sor.goods_value * 100):.1f}%"
        
        row = [
            str(index),  # Serial number
            str(sor.id),  # Original SOR ID
            sor.get_source_type_display(),
            str(sor.goods_value),
            sor.created_at.strftime('%d/%m/%Y') if sor.created_at else '--',
            sor.from_location[:15] + '...' if len(sor.from_location) > 15 else sor.from_location,
            sor.to_location[:15] + '...' if len(sor.to_location) > 15 else sor.to_location,
            (str(sor.vehicle)[:12] + '...' if sor.vehicle and len(str(sor.vehicle)) > 12 else str(sor.vehicle)) if sor.vehicle else (sor.outsourced_vehicle_text[:12] + '...' if sor.outsourced_vehicle_text and len(sor.outsourced_vehicle_text) > 12 else (sor.outsourced_vehicle_text or '--')),
            str(sor.vehicle.rate_per_km) if sor.vehicle and sor.vehicle.rate_per_km is not None else '--',
            f"{sor.distance_km:.1f}" if sor.distance_km else '--',
            transport_cost or '--',
            transport_percentage or '--',
            (str(sor.driver)[:12] + '...' if sor.driver and len(str(sor.driver)) > 12 else str(sor.driver)) if sor.driver else (sor.outsourced_driver_text[:12] + '...' if sor.outsourced_driver_text and len(sor.outsourced_driver_text) > 12 else (sor.outsourced_driver_text or '--')),
            sor.get_status_display()[:8] + '...' if len(sor.get_status_display()) > 8 else sor.get_status_display()
        ]
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response
